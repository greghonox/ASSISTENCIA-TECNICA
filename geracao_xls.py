from openpyxl import Workbook
from openpyxl.chart import ( PieChart3D, Reference)
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, colors, Border, Side, PatternFill, Alignment
from datetime import datetime

from tempfile import gettempdir
import sql

from PIL import Image as Imageresize

class Geracao_XLS:
    def __init__(self, caminho, dados, titulo):
        self.titulo = titulo
        self.dados = dados
        arquivo = Workbook()
        self.defini_formatacao()

        self.arquivo = Workbook()
        self.planilha = arquivo.active
        self.planilha.sheet_view.showGridLines = False
        self.planilha.sheet_view.zoomScale = 130

        if(titulo == 'CATEGORIAS'): self.tabela_categoria()
        elif(titulo == 'CLIENTES'): self.tabela_clientes()
        elif(titulo == 'FORNECEDOR' or titulo == 'TECNICOS'): self.tabela_for_tec()
        elif(titulo == 'USUARIOS'): self.tabela_usuario()
        elif(titulo == 'PRODUTOS'): self.tabela_produtos()
        elif(titulo == 'SERVIÇOS(MÃO DE OBRAS)'): self.tabela_servicos()
        elif(titulo == 'ESTOQUE'): self.tabela_estoque()

        self.set_border(self.planilha, 'A1:N' + str(self.pg_linha))
        self.main()
        arquivo.save(caminho)

    def set_border(self, ws, cell_range):
        rows = ws[cell_range]
        side = Side(border_style='thin', color="FF000000")

        rows = list(rows)
        max_y = len(rows) - 1  # index of the last row
        for pos_y, cells in enumerate(rows):
            max_x = len(cells) - 1  # index of the last cell
            for pos_x, cell in enumerate(cells):
                border = Border(
                    left=cell.border.left,
                    right=cell.border.right,
                    top=cell.border.top,
                    bottom=cell.border.bottom
                )
                if pos_x == 0:
                    border.left = side
                if pos_x == max_x:
                    border.right = side
                if pos_y == 0:
                    border.top = side
                if pos_y == max_y:
                    border.bottom = side

                # set new border only if it's one of the edge cells
                if pos_x == 0 or pos_x == max_x or pos_y == 0 or pos_y == max_y:
                    cell.border = border

    def defini_formatacao(self):
        borda = Side(style='medium', color=colors.BLACK)
        self.borda_definida = Border(left=borda, top=borda, bottom=borda, right=borda)
        self.cor_borda = PatternFill('solid', fgColor="e52d09")
        self.cor_titulo = PatternFill('solid', fgColor="C0C0C0")
        self.cor_corpo = PatternFill('solid', fgColor="808080")
        self.cor_sim = PatternFill('solid', fgColor="C0C0C0")
        self.cor_nao = PatternFill('solid', fgColor="ffffff")


    def main(self):
        self.criar_tabela()
        self.inserir_imagem()


    def criar_tabela(self):
        self.planilha.title = self.titulo
        self.arquivo.create_sheet(self.titulo)
        self.planilha.add_image(self.inserir_imagem(), "A1")


    def criar_planilha(self, arquivo):
        arquivo.save(self.caminho_arquivo)


    def inserir_imagem(self):
        try:
            if(self.gravar_img()):
                if(self.modificar_tamanho_img()):
                    return Image(gettempdir() + "/logo.png")
        except Exception as erro: print("ERRO EM PEGAR IMAGEM({})".format(erro)); return False


    def gravar_img(self):
        try:
            img =  sql.SQL().pegar_meus_dados_logo()[0][0]
            with open(gettempdir() + r'\logo.png', 'wb') as arq: arq.write(img)
            return True
        except Exception as erro: print("ERRO EM PEGAR IMAGEM DO BANCO ({})".format(erro)); return False


    def defini_contato(self):
        center_h_v = Alignment(horizontal='center', vertical='center', wrapText=True)
        self.planilha['C1'] = 'RELATÓRIO DE ' + self.titulo
        self.planilha["C1"].font = Font(size=30, bold="True")
        self.planilha["C1"].fill = self.cor_titulo
        self.planilha["C1"].border = self.borda_definida
        self.planilha["C1"].alignment = center_h_v
        self.planilha.merge_cells('C1:N3')

        self.planilha["A6"] = str("DATA DE EMISSÃO:"+
        str(datetime.now().day)+"/"+str(datetime.now().month)+"/"+
        str(datetime.now().year)+" \nHORA:"+str(datetime.now().hour)+":"+
        str(datetime.now().minute))
        self.planilha.merge_cells('A6:D7')
        self.planilha["A6"].alignment = center_h_v
        self.planilha['A6'].fill = self.cor_corpo

        for coluna in range(0, 4):
            for linha in range(0, 2):
                self.planilha[chr(65 + coluna) + str(linha + 6)].border = self.borda_definida

        dados = sql.SQL().pegar_meus_dados()[0]
        self.planilha["F6"] = 'EMPRESA: ' + dados[10] #EMPRESA
        self.planilha["F6"].alignment = center_h_v
        self.planilha.merge_cells("F6:N6")
        endereco = "ENDEREÇO:{} Nº {}\n {} {}-{} CEP:{}".format(dados[4], dados[6], dados[8], dados[7], dados[5], dados[3])
        self.planilha["F7"] = endereco.upper()# ENDEREÇO
        self.planilha["F7"].alignment = center_h_v
        self.planilha["F7"].fill = self.cor_titulo
        self.planilha.merge_cells("F7:N8")

        self.planilha["K9"] = dados[1] #TELEFONE
        self.planilha["K9"].alignment = center_h_v
        self.planilha.merge_cells("K9:N9")

        self.planilha["F9"] = dados[2]# e-mail
        self.planilha["F9"].alignment = center_h_v
        self.planilha.merge_cells("F9:J9")
        self.linha_dados = '12'

        for coluna in range(0, 9):
            for linha in range(0, 4):
                self.planilha[chr(70 + coluna) + str(linha + 6)].border = self.borda_definida


    def modificar_tamanho_img(self):
        try:
            altura = 120
            largura = 100
            img = Imageresize.open(gettempdir() + r'\logo.png')
            img = img.resize((altura, largura), Imageresize.ANTIALIAS)
            img.save(gettempdir() + r'\logo.png')
            return True
        except Exception as erro: print("ERRO EM MUDAR TAMANHO IMG({})".format(erro)); return False

    colunas = ['A', 'C', 'E', 'G', 'I', 'K', 'M']
    mesc = {'A': "A{}:B{}", 'C': "C{}:D{}", 'E': "E{}:F{}",
    'G': "G{}:H{}", 'I': "I{}:J{}", 'K': "K{}:L{}", 'M': "M{}:N{}"}
    def tabela_categoria(self):
        self.defini_contato()

        linha_d = int(self.linha_dados) + 1

        self.planilha['A'+ self.linha_dados] = 'CÓDIGO'
        self.planilha['C'+ self.linha_dados] = 'CATEGÓRIA'
        self.planilha['E'+ self.linha_dados] = 'TIPO CATEGORIA'
        self.planilha['G'+ self.linha_dados] = 'SITUAÇÃO'

        for col in (range(8)): self.planilha[chr(65 + col) + self.linha_dados].fill = self.cor_corpo
        [self.planilha.merge_cells(str(self.mesc[col]).format(self.linha_dados, self.linha_dados)) for col in self.colunas[0:4]]

        linha_sim = True
        for col, dados in enumerate(self.dados):
            for lin, dados_n in enumerate(dados):
                posicao = self.colunas[lin] + str(col + linha_d)
                self.planilha[posicao] = dados_n
                self.planilha.merge_cells(str(self.mesc[self.colunas[lin]]).format(col + linha_d, col + linha_d))
                self.planilha[posicao].fill = self.cor_sim if(linha_sim) else self.cor_nao
            linha_sim = not linha_sim
        self.pg_linha = col + int(self.linha_dados) + 1


    def tabela_clientes(self):
        self.defini_contato()

        linha_d = int(self.linha_dados) + 1

        self.planilha['A'+ self.linha_dados] = 'CÓDIGO'
        self.planilha['C'+ self.linha_dados] = 'CLIENTE'
        self.planilha['E'+ self.linha_dados] = 'TELEFONE'
        self.planilha['G'+ self.linha_dados] = 'WHATSAPP'
        self.planilha['I'+ self.linha_dados] = 'E-MAIL'
        self.planilha['K'+ self.linha_dados] = 'DOCUMENTO'
        self.planilha['M'+ self.linha_dados] = 'SITUACAO'

        for col in (range(14)): self.planilha[chr(65 + col) + self.linha_dados].fill = self.cor_corpo
        [self.planilha.merge_cells(str(self.mesc[col]).format(self.linha_dados, self.linha_dados)) for col in self.colunas[0:6]]

        linha_sim = True
        for col, dados in enumerate(self.dados):
            for lin, dados_n in enumerate(dados):
                posicao = self.colunas[lin] + str(col + linha_d)
                self.planilha[posicao] = dados_n
                self.planilha.merge_cells(str(self.mesc[self.colunas[lin]]).format(col + linha_d, col + linha_d))
                self.planilha[posicao].fill = self.cor_sim if(linha_sim) else self.cor_nao
            linha_sim = not linha_sim
        self.pg_linha = col + int(self.linha_dados) + 1


    def tabela_for_tec(self):
        self.defini_contato()
        linha_d = int(self.linha_dados) + 1

        self.planilha['A'+ self.linha_dados] = 'CÓDIGO'
        self.planilha['C'+ self.linha_dados] = 'PESSOA'
        self.planilha['E'+ self.linha_dados] = 'TELEFONE'
        self.planilha['G'+ self.linha_dados] = 'WHATSAPP'
        self.planilha['I'+ self.linha_dados] = 'E-MAIL'
        self.planilha['K'+ self.linha_dados] = 'DOCUMENTO'
        self.planilha['M'+ self.linha_dados] = 'SITUACAO'

        for col in (range(14)): self.planilha[chr(65 + col) + self.linha_dados].fill = self.cor_corpo
        [self.planilha.merge_cells(str(self.mesc[col]).format(self.linha_dados, self.linha_dados)) for col in self.colunas[0:6]]

        linha_sim = True
        for col, dados in enumerate(self.dados):
            for lin, dados_n in enumerate(dados):
                posicao = self.colunas[lin] + str(col + linha_d)
                self.planilha[posicao] = dados_n
                self.planilha.merge_cells(str(self.mesc[self.colunas[lin]]).format(col + linha_d, col + linha_d))
                self.planilha[posicao].fill = self.cor_sim if(linha_sim) else self.cor_nao
            linha_sim = not linha_sim
        self.pg_linha = col + int(self.linha_dados) + 1


    def tabela_usuario(self):
        self.defini_contato()
        linha_d = int(self.linha_dados) + 1

        self.planilha['A'+ self.linha_dados] = 'CÓDIGO'
        self.planilha['C'+ self.linha_dados] = 'USUÁRIO'
        self.planilha['E'+ self.linha_dados] = 'TELEFONE'
        self.planilha['G'+ self.linha_dados] = 'WHATSAPP'
        self.planilha['I'+ self.linha_dados] = 'E-MAIL'
        self.planilha['K'+ self.linha_dados] = 'SITUACAO'

        for col in (range(14)): self.planilha[chr(65 + col) + self.linha_dados].fill = self.cor_corpo
        [self.planilha.merge_cells(str(self.mesc[col]).format(self.linha_dados, self.linha_dados)) for col in self.colunas[0:5]]

        linha_sim = True
        for col, dados in enumerate(self.dados):
            for lin, dados_n in enumerate(dados):
                posicao = self.colunas[lin] + str(col + linha_d)
                self.planilha[posicao] = dados_n
                self.planilha.merge_cells(str(self.mesc[self.colunas[lin]]).format(col + linha_d, col + linha_d))
                self.planilha[posicao].fill = self.cor_sim if(linha_sim) else self.cor_nao
            linha_sim = not linha_sim
        self.pg_linha = col + int(self.linha_dados) + 1


    def tabela_produtos(self):
        self.defini_contato()

        linha_d = int(self.linha_dados) + 1

        self.planilha['A'+ self.linha_dados] = 'CÓDIGO'
        self.planilha['C'+ self.linha_dados] = 'PRODUTO'
        self.planilha['E'+ self.linha_dados] = 'FORNECEDOR'
        self.planilha['G'+ self.linha_dados] = 'UNIDADE'
        self.planilha['I'+ self.linha_dados] = 'CATEGORIA'
        self.planilha['K'+ self.linha_dados] = 'SITUAÇÃO'

        for col in (range(14)): self.planilha[chr(65 + col) + self.linha_dados].fill = self.cor_corpo
        [self.planilha.merge_cells(str(self.mesc[col]).format(self.linha_dados, self.linha_dados)) for col in self.colunas[0:6]]

        linha_sim = True
        for col, dados in enumerate(self.dados):
            for lin, dados_n in enumerate(dados):
                posicao = self.colunas[lin] + str(col + linha_d)
                self.planilha[posicao] = dados_n
                self.planilha.merge_cells(str(self.mesc[self.colunas[lin]]).format(col + linha_d, col + linha_d))
                self.planilha[posicao].fill = self.cor_sim if(linha_sim) else self.cor_nao
            linha_sim = not linha_sim
        self.pg_linha = col + int(self.linha_dados) + 1


    def tabela_servicos(self):
        self.defini_contato()

        linha_d = int(self.linha_dados) + 1

        self.planilha['A'+ self.linha_dados] = 'CÓDIGO'
        self.planilha['C'+ self.linha_dados] = 'SERVIÇO'
        self.planilha['E'+ self.linha_dados] = 'F.COBRANÇA'
        self.planilha['G'+ self.linha_dados] = 'CATEGORIA'
        self.planilha['I'+ self.linha_dados] = 'CUSTO'
        self.planilha['K'+ self.linha_dados] = 'SITUAÇÃO'

        for col in (range(14)): self.planilha[chr(65 + col) + self.linha_dados].fill = self.cor_corpo
        [self.planilha.merge_cells(str(self.mesc[col]).format(self.linha_dados, self.linha_dados)) for col in self.colunas[0:6]]

        linha_sim = True
        for col, dados in enumerate(self.dados):
            for lin, dados_n in enumerate(dados):
                posicao = self.colunas[lin] + str(col + linha_d)
                self.planilha[posicao] = dados_n
                self.planilha.merge_cells(str(self.mesc[self.colunas[lin]]).format(col + linha_d, col + linha_d))
                self.planilha[posicao].fill = self.cor_sim if(linha_sim) else self.cor_nao
            linha_sim = not linha_sim
        self.pg_linha = col + int(self.linha_dados) + 1


    def tabela_estoque(self):
        self.defini_contato()

        linha_d = int(self.linha_dados) + 1

        self.planilha['A'+ self.linha_dados] = 'CÓDIGO'
        self.planilha['C'+ self.linha_dados] = 'PRODUTO'
        self.planilha['E'+ self.linha_dados] = 'QUANTIDADE'
        self.planilha['G'+ self.linha_dados] = 'QUANTIDADE_MINIMA'
        self.planilha['I'+ self.linha_dados] = 'PREÇO'

        for col in (range(14)): self.planilha[chr(65 + col) + self.linha_dados].fill = self.cor_corpo
        [self.planilha.merge_cells(str(self.mesc[col]).format(self.linha_dados, self.linha_dados)) for col in self.colunas[0:6]]

        linha_sim = True
        for col, dados in enumerate(self.dados):
            for lin, dados_n in enumerate(dados):
                posicao = self.colunas[lin] + str(col + linha_d)
                self.planilha[posicao] = dados_n
                self.planilha.merge_cells(str(self.mesc[self.colunas[lin]]).format(col + linha_d, col + linha_d))
                self.planilha[posicao].fill = self.cor_sim if(linha_sim) else self.cor_nao
            linha_sim = not linha_sim
        self.pg_linha = col + int(self.linha_dados) + 1


# Geracao_XLS(r'C:\Users\SUPORTE05\Desktop\doc.xlsx',
#              [['1', 'PRODUTO DE REDES', '0', '0'],
#              ['2', 'MANUNTECAO DE REDES', '1', '1'],
#              ['3', 'CONFIGURAÇÃO DE COMPUTADORES','1','1'],
#              ['4', 'MANUTENÇÃO DE COMPUTADORE','1','1'],
#              ['5', 'IMPLANTAÇÃO DE ESTABELECIMENTO','1','1'],
#              ['6', 'SUPLEMENTO DE INFORMATICA','0','1'],
#              ['7', 'CONFIGURAÇÃO DE REDES','1','1'],
#              ['8', 'FERRAMENTAS', '0',	'1']],
#              "CATEGORIAS")
